from flask import Flask, request, jsonify, session, redirect, send_file
from flask_cors import CORS 
from werkzeug.utils import secure_filename
import os
import mysql.connector
from mysql.connector import Error
import openpyxl
import pandas as pd 
from io import BytesIO
import requests
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import traceback
import logging
from datetime import datetime
import bcrypt

app = Flask(__name__)
app.config['SECRET_KEY'] = "enmasse4ever"

CORS(app, resources={r"/*": {"origins": "*"}})

DATA_QUERY_URL = "http://localhost:5000/upload"  # URL of the data_query Flask app

# Database configuration
db_config = {
    'host': 'localhost',
    'user': 'root',
    'password': '',  # Enter your MySQL password here
    'database': 'enmasse',
}

def execute_sql_file(filename, cursor):
    with open(filename, 'r') as file:
        sql = file.read()
    for result in cursor.execute(sql, multi=True):
        pass

def create_unique_table_name(username, cursor):
    suffix = 1
    base_table_name = f"{username}_data"
    table_name = base_table_name

    cursor.execute("SHOW TABLES")
    existing_tables = cursor.fetchall()
    existing_table_names = [table[0] for table in existing_tables]

    while table_name in existing_table_names:
        table_name = f"{base_table_name}_{suffix}"
        suffix += 1

    return table_name

@app.route('/login', methods=['POST'])
def login():
    try:
        if request.is_json:
            data = request.get_json()
            login_id = data.get('login')
            password = data.get('password')
        else:
            login_id = request.form.get('login')
            password = request.form.get('password')

        if not login_id or not password:
            raise ValueError("Missing login or password in request")

        db = mysql.connector.connect(**db_config)
        cursor = db.cursor(dictionary=True)

        query = "SELECT * FROM login WHERE username = %s OR email = %s"
        cursor.execute(query, (login_id, login_id))
        user = cursor.fetchone()

        if user and 'password' in user and bcrypt.checkpw(password.encode('utf-8'), user['password'].encode('utf-8')):
            session['username'] = user['username']

            # Record login date
            login_time = datetime.now()
            query = "INSERT INTO user_sessions (user_id, login_date) VALUES (%s, %s)"
            cursor.execute(query, (user['user_id'], login_time))
            db.commit()

            cursor.close()
            db.close()

            return jsonify({"message": "Login successful", "user": user}), 200
        else:
            cursor.close()
            db.close()
            return jsonify({"message": "Invalid username/email or password"}), 401
    except (Error, ValueError) as e:
        return jsonify({"message": "An error occurred", "error": str(e)}), 500

@app.route('/upload_page', methods=['POST'])
def upload_page():
    if 'username' not in request.form:
        return jsonify({"message": "Username not provided"}), 400

    username = request.form['username']

    if 'file' not in request.files:
        return jsonify({"message": "No file part"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"message": "No selected file"}), 400

    if file:
        try:
            # Validate file format
            file_extension = file.filename.split('.')[-1].lower()
            if file_extension != 'xlsx':
                return jsonify({"message": "Invalid file format. Only .xlsx files are allowed."}), 400

            # Validate file structure
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            headers = [cell.value.lower() for cell in sheet[1]]

            if len(headers) < 2 or 'name' not in headers or 'email' not in headers:
                return jsonify({"message": "Invalid file structure. Excel file must contain 'name' and 'email' columns."}), 400

            # Process the valid file
            db = mysql.connector.connect(**db_config)
            cursor = db.cursor()

            # Get user ID from the login table
            cursor.execute("SELECT user_id FROM login WHERE username = %s", (username,))
            user = cursor.fetchone()
            if not user:
                return jsonify({"message": "User not found"}), 400
            user_id = user[0]

            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)

            table_name = create_unique_table_name(username, cursor)

            cursor.execute(f"""
            CREATE TABLE {table_name} (
                id INT AUTO_INCREMENT PRIMARY KEY,
                name VARCHAR(255) NOT NULL,
                email VARCHAR(255) NOT NULL,
                INDEX(name),
                INDEX(email)
            )
            """)

            for row in data:
                if len(row) >= 2:
                    name, email = row[0], row[1]
                    cursor.execute(f"INSERT INTO {table_name} (name, email) VALUES (%s, %s)", (name, email))

            # Insert information into user_uploaded_tables
            cursor.execute("INSERT INTO user_uploaded_tables (user_id, table_name) VALUES (%s, %s)", (user_id, table_name))

            db.commit()
            cursor.close()
            db.close()

            return jsonify({"message": "File uploaded and processed successfully"}), 200
        except Exception as e:
            return jsonify({"message": "Error processing the file", "error": str(e)}), 500

@app.route('/logout', methods=['GET'])
def logout():
    try:
        # Clear the session
        session.pop('username', None)
        
        # Redirect to Flask logout endpoint
        return redirect("/login")
    except Exception as e:
        # Log any errors and handle them appropriately
        print(f"Error during logout: {e}")
        return redirect("/login")  # Redirect to login page even if an error occurs

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'username' not in request.form:
        return jsonify({"message": "Username not provided"}), 400

    username = request.form['username']

    if 'file' not in request.files:
        return jsonify({"message": "No file part"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"message": "No selected file"}), 400

    if file:
        try:
            db = mysql.connector.connect(**db_config)
            cursor = db.cursor()

            # Get user ID from the login table
            cursor.execute("SELECT user_id FROM login WHERE username = %s", (username,))
            user = cursor.fetchone()
            if not user:
                return jsonify({"message": "User not found"}), 400
            user_id = user[0]

            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)

            table_name = create_unique_table_name(username, cursor)

            cursor.execute(f"""
            CREATE TABLE {table_name} (
                id INT AUTO_INCREMENT PRIMARY KEY,
                name VARCHAR(255) NOT NULL,
                email VARCHAR(255) NOT NULL,
                INDEX(name),
                INDEX(email)
            )
            """)

            for row in data:
                if len(row) >= 2:
                    name, email = row[0], row[1]
                    cursor.execute(f"INSERT INTO {table_name} (name, email) VALUES (%s, %s)", (name, email))

            # Insert information into user_uploaded_tables
            cursor.execute("INSERT INTO user_uploaded_tables (user_id, table_name) VALUES (%s, %s)", (user_id, table_name))

            # Record file upload metrics
            file_type = file.content_type
            file_size = len(file.read())
            cursor.execute("INSERT INTO file_uploads (user_id, file_name, success, file_type, file_size) VALUES (%s, %s, %s, %s, %s)",
                           (user_id, file.filename, True, file_type, file_size))

            # Update or insert into data_table_metrics
            cursor.execute("SELECT * FROM data_table_metrics WHERE user_id = %s", (user_id,))
            metrics = cursor.fetchone()
            if metrics:
                cursor.execute("UPDATE data_table_metrics SET tables_created = tables_created + 1 WHERE user_id = %s", (user_id,))
            else:
                cursor.execute("INSERT INTO data_table_metrics (user_id, tables_created, tables_deleted) VALUES (%s, %s, %s)", (user_id, 1, 0))

            db.commit()
            cursor.close()
            db.close()

            return jsonify({"message": "File uploaded and processed successfully"}), 200
        except Exception as e:
            return jsonify({"message": "Error processing the file", "error": str(e)}), 500



@app.route('/get_user_tables', methods=['POST'])
def get_user_tables():
    data = request.get_json()
    username = data.get('username')

    if not username:
        return jsonify({"message": "Username not provided"}), 400

    try:
        db = mysql.connector.connect(**db_config)
        cursor = db.cursor()

        cursor.execute("""
            SELECT table_name 
            FROM user_uploaded_tables 
            WHERE user_id = (SELECT user_id FROM login WHERE username = %s)
        """, (username,))
        results = cursor.fetchall()
        cursor.close()
        db.close()

        tables = [row[0] for row in results]

        return jsonify({"message": "Success", "tables": tables}), 200
    except mysql.connector.Error as err:
        return jsonify({"message": "An error occurred", "error": str(err)}), 500

@app.route('/fetch_table_data', methods=['POST'])
def fetch_table_data():
    data = request.json  # Access JSON data sent from the frontend

    if 'table_name' not in data:
        return jsonify({"message": "Table name not provided"}), 400

    table_name = data['table_name']  # Extract table name from JSON payload

    try:
        db = mysql.connector.connect(**db_config)
        cursor = db.cursor(dictionary=True)

        cursor.execute(f"SELECT name, email FROM {table_name}")
        table_data = cursor.fetchall()

        cursor.close()
        db.close()

        return jsonify({"table_data": table_data}), 200
    except mysql.connector.Error as err:
        return jsonify({"message": "An error occurred", "error": str(err)}), 500

@app.route('/export_table', methods=['GET'])
def export_table():
    table_name = request.args.get('table_name')
    if not table_name:
        return jsonify({"message": "Table name not provided"}), 400

    try:
        db = mysql.connector.connect(**db_config)
        cursor = db.cursor()

        # Fetch table data excluding the 'id' column
        cursor.execute(f"SELECT name, email FROM {table_name}")
        table_data = cursor.fetchall()

        # Create a DataFrame
        df = pd.DataFrame(table_data)

        # Prepare the Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, header=False, sheet_name='Sheet1')

        output.seek(0)
        cursor.close()
        db.close()

        # Send the Excel file
        return send_file(output, download_name=f"{table_name}.xlsx", as_attachment=True)

    except mysql.connector.Error as err:
        return jsonify({"message": "An error occurred", "error": str(err)}), 500

@app.route('/delete_table', methods=['POST'])
def delete_table():
    data = request.get_json()
    table_name = data.get('table_name')
    username = data.get('username')

    if not table_name or not username:
        return jsonify({"message": "Table name or username not provided"}), 400

    try:
        db = mysql.connector.connect(**db_config)
        cursor = db.cursor()

        # Fetch user_id from the username
        cursor.execute("SELECT user_id FROM login WHERE username = %s", (username,))
        user = cursor.fetchone()

        if not user:
            return jsonify({"message": "User not found"}), 404

        user_id = user[0]  # Extract user_id from the result

        # Delete the table
        cursor.execute(f"DROP TABLE IF EXISTS {table_name}")

        # Delete the entry from user_uploaded_tables
        cursor.execute("DELETE FROM user_uploaded_tables WHERE table_name = %s AND user_id = %s", (table_name, user_id))

        # Update or insert into data_table_metrics
        cursor.execute("SELECT * FROM data_table_metrics WHERE user_id = %s", (user_id,))
        metrics = cursor.fetchone()
        if metrics:
            cursor.execute("UPDATE data_table_metrics SET tables_deleted = tables_deleted + 1 WHERE user_id = %s", (user_id,))
        else:
            cursor.execute("INSERT INTO data_table_metrics (user_id, tables_created, tables_deleted) VALUES (%s, %s, %s)", (user_id, 0, 1))

        db.commit()
        cursor.close()
        db.close()

        return jsonify({"message": "Table and entry deleted successfully"})

    except mysql.connector.Error as err:
        return jsonify({"message": "An error occurred", "error": str(err)}), 500
    
@app.route('/save_table', methods=['POST'])
def save_table():
    data = request.get_json()
    table_name = data.get('table_name')
    table_data = data.get('table_data')
    username = data.get('username')

    if not table_name or not table_data or not username:
        return jsonify({"message": "Table name, data, or username not provided"}), 400

    try:
        db = mysql.connector.connect(**db_config)
        cursor = db.cursor()

        # Fetch user_id from the username
        cursor.execute("SELECT user_id FROM login WHERE username = %s", (username,))
        user_id = cursor.fetchone()

        if not user_id:
            return jsonify({"message": "User not found"}), 404

        # Drop the existing table
        cursor.execute(f"DROP TABLE IF EXISTS {table_name}")

        # Create the table again with the same structure (assuming name and email columns)
        cursor.execute(f"CREATE TABLE {table_name} (name VARCHAR(255), email VARCHAR(255))")

        # Insert new data into the table
        for row in table_data:
            cursor.execute(f"INSERT INTO {table_name} (name, email) VALUES (%s, %s)", (row['name'], row['email']))

        db.commit()
        cursor.close()
        db.close()

        return jsonify({"message": "Table data saved successfully"})

    except mysql.connector.Error as err:
        return jsonify({"message": "An error occurred", "error": str(err)}), 500


def send_email(sender_email, password, recipients, subject, message):
    try:
        smtp_server = 'smtp.mailgun.org'
        smtp_port = 587

        # Prepare the email message
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = ", ".join(recipients)
        msg['Subject'] = subject
        msg.attach(MIMEText(message, 'plain'))

        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()  # Secure the connection
            server.login(sender_email, password)
            server.sendmail(sender_email, recipients, msg.as_string())

        print('Email sent successfully!')

    except smtplib.SMTPAuthenticationError as e:
        error_message = "SMTP Authentication Error: " + str(e)
        logging.error(error_message)
        raise e
    except Exception as e:
        error_message = "Error sending email: " + str(e)
        logging.error(error_message)
        traceback.print_exc()  # Log the full traceback for debugging
        raise e

def send_email_attachment(sender_email, password, recipients, subject, message, attachments=None):
    try:
        # Outlook SMTP server details
        smtp_server = 'smtp.mailgun.org'
        smtp_port = 587

        # Connect to the Outlook SMTP server
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()  # Secure the connection
            server.login(sender_email, password)

            # Iterate over each recipient and their corresponding attachment
            for recipient, attachment in zip(recipients, attachments or []):
                # Prepare the email message
                msg = MIMEMultipart()
                msg['From'] = sender_email
                msg['To'] = recipient
                msg['Subject'] = subject
                msg.attach(MIMEText(message, 'plain'))

                # Attach the file if provided
                if attachment:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(attachment.read())
                    encoders.encode_base64(part)
                    part.add_header('Content-Disposition', f"attachment; filename={secure_filename(attachment.filename)}")
                    msg.attach(part)

                # Send the email to the current recipient
                server.sendmail(sender_email, recipient, msg.as_string())

        print('Emails sent successfully!')

    except smtplib.SMTPAuthenticationError as e:
        error_message = "SMTP Authentication Error: " + str(e)
        logging.error(error_message)
        raise e
    except Exception as e:
        error_message = "Error sending emails: " + str(e)
        logging.error(error_message)
        traceback.print_exc()  # Log the full traceback for debugging
        raise e

@app.route('/send-email', methods=['POST'])
def send_email_route():
    try:
        # Get form data
        recipients = request.form.getlist('recipientsEmail[]')
        subject = request.form['subject']
        message = request.form['message']
        attachments = request.files.getlist('attachmentsInput[]')
        sender_email = request.form['email']  # Retrieve sender email from form data

        # Collect SMTP credentials
        password = request.form['password']
        
        if not attachments:
            send_email(sender_email, password, recipients, subject, message)
            return jsonify({"message": "Emails sent successfully"}), 200
        else:
            send_email_attachment(sender_email, password, recipients, subject, message, attachments)
            return jsonify({"message": "Emails sent successfully"}), 200
            
    except smtplib.SMTPAuthenticationError as e:
        error_message = "SMTP Authentication Error: " + str(e)
        logging.error(error_message)
        return jsonify({"message": "Failed to authenticate with SMTP server", "error": error_message}), 500
    except Exception as e:
        error_message = "Error sending emails: " + str(e)
        logging.error(error_message)
        return jsonify({"message": "Failed to send emails", "error": error_message}), 500


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)